import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

FILE = "scores.xlsx"

def create_and_write():
    """文件不存在时创建Excel并写入数据"""
    # 判断文件是否存在
    if os.path.exists(FILE):
        raise FileExistsError(f"文件 {FILE} 已存在，无需创建")

    # 新建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩表"

    # 写入表头
    headers = ["姓名", "语文", "数学"]
    ws.append(headers)

    # 写入数据行（append 自动逐行追加）
    data = [
        ("张三", 88, 92),
        ("李四", 55, 60),
        ("王五", 95, 90),
    ]
    for row in data:
        ws.append(row)

    wb.save(FILE)
    wb.close()
    print(f"已创建并写入 {FILE}")


def write_result_and_read():
    """读取已有文件，计算并写回结果，最后读取验证"""
    # 加载已有文件（文件不存在会抛 FileNotFoundError）
    wb = load_workbook(FILE)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # 追加“总分 / 平均分 / 等级”三列表头
    title_font = Font(bold=True)
    center = Alignment(horizontal="center")
    for i, col in enumerate(["总分", "平均分", "等级"], start=1):
        cell = ws.cell(row=1, column=max_col + i)
        cell.value = col
        cell.font = title_font
        cell.alignment = center

    # 逐行计算并写入
    for row in range(2, max_row + 1):
        chinese = ws.cell(row=row, column=2).value
        math = ws.cell(row=row, column=3).value

        # 校验数据，防止 None 或非数字导致计算错误
        if chinese is None or math is None:
            raise ValueError(f"第{row}行存在空值，无法计算")
        if not isinstance(chinese, (int, float)) or not isinstance(math, (int, float)):
            raise TypeError(f"第{row}行成绩不是数字：语文={chinese}, 数学={math}")

        total = chinese + math
        avg = total / 2

        # 写入总分、平均分、等级
        ws.cell(row=row, column=max_col + 1, value=total).alignment = center
        ws.cell(row=row, column=max_col + 2, value=avg).alignment = center
        ws.cell(row=row, column=max_col + 3, value="优秀" if avg >= 90 else "良好" if avg >= 80 else "及格" if avg >= 60 else "不及格").alignment = center

    wb.save(FILE)
    wb.close()
    print(f"计算完成，结果已写回 {FILE}")

    # ===== 后置步骤：读取验证写入结果 =====
    verify_wb = load_workbook(FILE)
    verify_ws = verify_wb.active
    print("\n读取验证结果：")
    for row in verify_ws.iter_rows(values_only=True):
        print(row)
    verify_wb.close()


def main():
    # 第一步：没有文件就先创建
    if not os.path.exists(FILE):
        create_and_write()
    # 第二步：读取 → 计算 → 写回 → 再读取
    write_result_and_read()


if __name__ == "__main__":
    main()
